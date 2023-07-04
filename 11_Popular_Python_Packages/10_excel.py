import openpyxl

# openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Tabelle1"]

# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

cell = sheet["a1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

sheet.cell(row=1, column=1)

print(sheet.max_row)
print(sheet.max_column)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

column = sheet["a"]
print(column)

cells = sheet["a:c"]
print(cells)

cells = sheet["a1:c3"]
print(sheet[1])  # print first row
print(sheet[1:3])  # print first to third row

sheet.append([1004, 4, 8.95])
wb.save("transactions2.xlsx")
